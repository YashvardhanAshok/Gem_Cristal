raw_text='''******
6X
A-Z OFFICE SUPPLIERS
A.K.TRADE AGENCY
A.V.M. SALES CORPORATION
AADRIKA AUTOMATION
AAN Associates
AASTHA CORNER
AGRAWAL ENTERPRISES
AJAY ENTERPRISE
AR ENTERPRISES
ARCHANA MEDICARI
ASHA MEDICAL STORE
Alfa Biomedical Co
Ankit vastralaya & readymade
Arbind Enterprise
B K THAPAR HOSIERY WORKS
BHATI BUSINESS SYNDICATE
BLUE SQUARE
Bharat Enterprise
Bharat Metals
CH JAGANNATH MADANLAL
CHANDAN KUMAR
CHANDRA OVERSEAS
CHANDRESWAR SINGH
COMPUTORIUM
CONTINENTAL MILKOSE
CRYSTAL WORKS
CSA CORPORATION PRIVATE LIMITED
D. K. ENTERPRISE
DEB VARIETIES
DEEPAK ENTERPRISE
DELVI TREXIM
DERPA INDUSTRIAL POLYMERS PVT LTD
DIAGNOSTICMED
Deep Systems & Solutions
ENTREMONDE POLYCOATERS LIMITED
ERA GLOBAL STANDARDS CERTIFICATION PRIVATE LIMITED
ESTEEM360 ZONE PRIVATE LIMITED
EWIT INFOTECH
EXICOM TECHNOLOGIES INDIA PRIVATE LIMITED
F I ENTERPRISES
FAIRDEAL HEALTH CARE PRIVATE LIMITED
FIRDOUS AND CO
FLORENCE INDIA
GARG ASSOCIATES
GARV ENTERPRISES
GOODWILL FIRM
GURJYOT ENTERPRISES
H.H.MARKETING
H.P.SINGH MACHINERY PVT LTD
HANS RAJ
HANUMANBUX UMADUTT
HARDIK ENTERPRISES
HARVINDER SINGH SEHGAL
HILLS STATIONERS & PRINTERS
HIRAMANI ENTERPRISES
ICON POWER SOLUTIONS PRIVATE LIMITED
ISLAND LALOO
J B TRADERS
JALARAM TREDARS
JE PREFAB INFRA PRIVATE LIMITED
K U ENTERPRISES
K-9 INDIA SOLUTIONS
K.L. EQUIPMENTS
KAP CONSTRUCTION
KRISHA BUSINESS LLP
KRITIKA MEDI PHARMA
LE MERITE EXPORTS LIMITED
LEELA TRADING COMPANY
LOGIX NET SOLUTIONS PRIVATE LIMITED
Lucky Enterprises
M.S.TRADERS
M/S B.INTERNATIONAL
M/S BHARAT KUMAR
M/S BIANCA MEDICOS
M/S CARON ENTERPRISES
M/S DEEPAK ENTERPRISES
M/S G. M. ENTERPRISES
M/S G. N. TRADERS
M/S HAKEEM MEDICAL HALL
M/S HARMEET KAUR SEHGAL
M/S KANISHKA ENTERPRISE
M/S KAUSHAL AGENCIES
M/S KEVICHOL SOPHIE
M/S KRITIKA INDUSTRISES
M/S LIPIKA ENTERPRISE,GOROIMARI,AHOM CHUBURI,SONITPUR.
M/S LOTHA TRADING CO.
M/S M.K. AUTOMOBILE
M/S M.K. TRADERS
M/S MAKMONI ENTERPRISES
M/S MAMTA PRODUCTS
M/S MB ENTERPRISES
M/S NAVSHEEN ASSOCIATES
M/S P ENTERPRISES
M/S P K ELECTRONICS
M/S RAHUL TRADING COMPANY
M/S RAM BABU YADAV
M/S RAMAWATAR SINGH
M/S SANJAYKUMAR SINGH
M/S SATURN ENTERPRISES
M/S SAWAI SINGH
M/S SHAKTI CONSTRUCTION & SUPPLY CO.
M/S SHEIKH ENTERPRISES
M/S SHIV SHAKTI DISTRIBUTORS
M/S SHREE SHANTI STATIONERY
M/S STANLEE MAHONGNAO
M/S SUCHI ENTERPRISES
M/S TOHOKHU T SEMA
M/S UMA ENTERPRISES
M/S YOGESH CHANDRA AGARWALA
M/S. GARG GENERAL STORE
M/S. HBM ENTERPRISE
M/S. K.K.ENTERPRISES
M/S. MODERN MACHINERY STORES
M/S. TANWAR TRADERS
M/S. UDAY KUMAR SINGH
M/S. UMRAO LAL GOYAL & SONS
M/s Harish Kumar Kushwaha
M/s Mahesh Kumar Gupta
M/s. Jamuna Enterprises
M/s.ZUHOLI SUMI
MAA DURGA ENTERPRISE
MAA GAYATRI DRUG DISTRIBUTORS
MAA SHAKAMBARI SUPPLIER
MAANVIK TRADING
MAHAJONG COKE LLP
MAHALAXMI ENTERPRISES
MANIFEST ENTERPRISES & CO
MANISHA TRADING COMPANY
MARUTI MARKETING
MD ABDUL MATIN
MEGHALAYA CONSTRUCTION & SUPPLY CO.
MICROBION LIFESCIENCE PRIVATE LIMITED
MIKADO ENGINEERS
ML SONS
MS BISHAKA JAIN
MS ENTERPRISES
Mangla Plastic Industries
Multiventure Group of companies
N E ENTERPRISES
NCS ENTERPRISES
NEELKANTH TRADER
NICE INFOTECH
NORTH EAST TRADE CENTRE
NR Solutions
NandAnand Co.
O.P ENGINEER'S
OMKAR STEEL WORKS
Office Kart
PADAMAVATI ENTERPRISES
PATEL MEDICARE
PINTU KUMAR JHA
R S TRADE & AGENCY PRIVATE LIMITED
RA SURGI PHARMA
RADHA TRADING COMPANY
RARA SUPPLY
RAUSHEENA UDYOG LIMITED
RESHI CONSTRUCTIONS AND SERVICES PRIVATELIMITED
RNY & SONS
RYAN ENTERPRISES
S M Enterprises
S V ENTERPRISE
S. G. ENGINEERS
S. TRADERS
SAMRAT INDER SINGH
SANCHAR WIRELESS COMMUNICATIONS LIMITED
SANDEEP KUMAR AGRAWAL
SATYAM TRADERS
SHARMA TRADER
SHREE SAI ENTERPRISES
SHRI STERON KHARJANA
SHRI SUBHASH AGARWALLA
SINGH AND COMPANY
SINHA ENTERPRISE
SOURCE DOT COM PRIVATE LIMTED
SPHERE COM SERVICES PRIVATE LIMITED
SREE RAM ASSOCIATES
SUGANDHA TRADING SOLUTIONS
SUNIL PHARMACY
SWADESHI WOOLLEN MILLS
Shrishti electrical
TINSUKIA TRADE AND SUPPLY
Tayal Electric & Construction Co
UNIQUE SERVICES PRIVATE LIMITED
UNIVERSAL PRODUCTS
V D INTELLISYS TECHNOLOGIES PRIVATE LIMITED
VEER TRADING COMPANY
VINAYAK ENTERPRISE
VISHAL ENTERPRISE
VISHWA TRADERS
lluvia tea


'''

import pyodbc
import pandas as pd
import ast
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

query = """
    SELECT * FROM tender_data 
    WHERE L_Placeholder IS NOT NULL 
    AND L_Placeholder != 'null' and organisation not in ('ASSAM RIFLES');
"""
df = pd.read_sql(query, conn)

parsed_rows = []

for _, row in df.iterrows():
    l_placeholder = row.get("L_Placeholder", "")
    try:
        placeholder_data = ast.literal_eval(l_placeholder)
        if isinstance(placeholder_data, list) and len(placeholder_data) > 0:
            entry = placeholder_data[0]  # Only take the first one (L1)
            if isinstance(entry, list) and len(entry) == 2:
                name_raw = entry[0]
                amount_str = entry[1].split()[0].replace(",", "")
                try:
                    amount = float(amount_str)
                except:
                    continue
                name_clean = name_raw.split('(')[0].strip()
                parsed_rows.append({
                    'Name': name_clean,
                    'Tender_id': row.get('tender_id', ''),
                    'qty': row.get('qty', ''),
                    'Item_description': row.get('item_description', ''),
                    'Department': row.get('organisation', ''),
                    'Start_date': row.get('start_date', ''),
                    'End_date': row.get('end_date', ''),
                    'Tender_value': row.get('tender_value', ''),
                    'Amount': amount,
                    'address': row.get('address', ''),
                    
                    'L_Label': 'L1'
                })
    except Exception:
        continue

if not parsed_rows:
    print("No L1 rows found.")
    exit()

final_df = pd.DataFrame(parsed_rows)

# Add SN and Count
grouped = final_df.groupby('Name').agg(Count=('Tender_id', 'count')).reset_index()
final_df = final_df.merge(grouped, on='Name', how='left')
final_df.insert(0, 'Sn', final_df.groupby('Name').ngroup() + 1)
final_df.sort_values(by=['Sn', 'Name', 'Tender_id'], inplace=True)

# Final columns
final_df = final_df[['Sn', 'Count', 'Name', 'L_Label', 'Tender_id', 'Department', 'Item_description', 'qty','address', 'Start_date', 'End_date', 'Tender_value', 'Amount']]

# Save to Excel
output_path = "Only_L1_Bidders_full_with_New_list.xlsx"
final_df.to_excel(output_path, index=False)

# Excel formatting (optional striping)
wb = load_workbook(output_path)
ws = wb.active

gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

current_sn = None
fill_toggle = False
for row in range(2, ws.max_row + 1):  
    sn_cell = ws.cell(row=row, column=1)
    sn_value = sn_cell.value

    if sn_value != current_sn:
        fill_toggle = not fill_toggle
        current_sn = sn_value

    fill = gray_fill if fill_toggle else white_fill
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill

wb.save(output_path)
print(f"Excel with only L1 bidders saved: {output_path}")
















































































# all 
# import pyodbc
# import pandas as pd
# import ast
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# conn = pyodbc.connect(
#     "DRIVER={ODBC Driver 17 for SQL Server};"
#     "SERVER=localhost\\SQLEXPRESS;"
#     "DATABASE=gem_tenders;"
#     "Trusted_Connection=yes;"
# )

# query = """
#     SELECT * FROM tender_data 
#     WHERE L_Placeholder IS NOT NULL 
#     AND L_Placeholder != 'null' and organisation in('ASSAM RIFLES')
# """
# df = pd.read_sql(query, conn)

# parsed_rows = []

# # Provide the company list you want to filter (optional)

# company_list = raw_text.lower().strip().split('\n')

# for _, row in df.iterrows():
#     l_placeholder = row.get("L_Placeholder", "")
#     try:
#         placeholder_data = ast.literal_eval(l_placeholder)
#         if isinstance(placeholder_data, list):
#             for idx, entry in enumerate(placeholder_data):
#                 if isinstance(entry, list) and len(entry) == 2:
#                     name_raw = entry[0]
#                     amount_str = entry[1].split()[0].replace(",", "")
#                     try:
#                         amount = float(amount_str)
#                     except:
#                         continue
#                     name_clean = name_raw.split('(')[0].strip()
#                     if name_clean.lower() in company_list:
#                         parsed_rows.append({
#                             'Name': name_clean,
#                             'Tender_id': row.get('tender_id', ''),
#                             'qty': row.get('qty', ''),
#                             'Item_description': row.get('item_description', ''),
#                             'Department': row.get('organisation', ''),
#                             'Start_date': row.get('start_date', ''),
#                             'End_date': row.get('end_date', ''),
#                             'Tender_value': row.get('tender_value', ''),
#                             'Amount': amount,
#                             'L_Label': f"L{idx+1}"  # L1, L2, L3 based on index
#                         })
#     except Exception:
#         continue

# if not parsed_rows:
#     print("No valid L_Placeholder data found.")
#     exit()

# parsed_df = pd.DataFrame(parsed_rows)

# # Add SN and Count
# grouped = parsed_df.groupby('Name').agg(Count=('Tender_id', 'count')).reset_index()
# final_df = parsed_df.merge(grouped, on='Name', how='left')
# final_df.insert(0, 'Sn', final_df.groupby('Name').ngroup() + 1)
# final_df.sort_values(by=['Sn', 'Name', 'Tender_id'], inplace=True)

# # Final column order
# final_df = final_df[['Sn', 'Count', 'Name', 'L_Label', 'Tender_id', 'Department', 'Item_description', 'qty', 'Start_date', 'End_date', 'Tender_value', 'Amount']]

# # Save to Excel
# output_path = "L_Labelled_By_Index_assam.xlsx"
# final_df.to_excel(output_path, index=False)

# # Excel styling
# wb = load_workbook(output_path)
# ws = wb.active

# gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
# white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# current_sn = None
# fill_toggle = False
# for row in range(2, ws.max_row + 1):  
#     sn_cell = ws.cell(row=row, column=1)
#     sn_value = sn_cell.value

#     if sn_value != current_sn:
#         fill_toggle = not fill_toggle
#         current_sn = sn_value

#     fill = gray_fill if fill_toggle else white_fill
#     for col in range(1, ws.max_column + 1):
#         ws.cell(row=row, column=col).fill = fill

# wb.save(output_path)
# print(f"Excel with L1/L2 based on list index saved: {output_path}")
