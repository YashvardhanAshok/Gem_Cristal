
# Keywords to filter and export




import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.page import PageMargins
import os

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

# Fetch data
query = "SELECT * FROM tender_data WHERE end_date > CAST(GETDATE() AS DATE) AND Cancel NOT IN ('Cancel')"
df = pd.read_sql(query, conn)

# Drop unnecessary columns
columns_to_drop = ['id', "element_put", 'branch', "item_category", "consignee_reporting", "date_of_search", "updated_at", 'file_path', 'link_href', 'Live', "extended", "L1_update", 'status', 'L_Placeholder', "Cancel","state","epbg_percentage","matches","ministry"]

df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True)

# Replace 0s with empty strings
df.replace(0, '', inplace=True)

# Convert tender value to words
def convert_to_words(val):
    try:
        val = float(val)
        if val >= 1_00_00_000:
            return f"{val / 1_00_00_000:.1f} Cr"
        elif val >= 1_00_000:
            return f"{val / 1_00_000:.1f} L"
        elif val > 0:
            return f"{val:.0f}"
        else:
            return ""
    except:
        return ""

# Add converted tender value
if 'tender_value' in df.columns:
    df['Ten-Val Word'] = df['tender_value'].apply(convert_to_words)

# Format column names
df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

# Move 'Ten-Val Word' next to 'Tender Valu'
cols = df.columns.tolist()
if 'Tender Valu' in cols and 'Ten-Val Word' in cols:
    tv_index = cols.index('Tender Valu')
    cols.remove('Ten-Val Word')
    cols.insert(tv_index + 1, 'Ten-Val Word')
    df = df[cols]

# Output Excel file path
save_file = os.path.abspath(os.path.join(os.path.dirname(__file__), "xl files"))
os.makedirs(save_file, exist_ok=True)
output_file = os.path.join(save_file, "Main_no-workbook.xlsx")

# Define keyword groups
keyword_groups = [
    ["prefab shelters with puf panel"],
    ["lgsf building"],
    ["inflatable shelters"],
    ["porta cabin"],
    ["portable houses"],
    ["portable kitchen"],
    ["ppgi sheets"],
    ["cgi sheet"],
    ["led bulbs"],
    ["street light"],
    ["flood light"],
    ["gyser"],
    ["room heater"],
    ["xlpe cables"],
    ["wire"],
    ["cables"],
    ["mcb"],
    ["mccb"],
    ["ac static watthour ", "meters-energy meter"],
    ["switch fuse unit"],
    ["decorative street light"],
    ["decorative bollard"],
    ["water cooler"],
    ["lighting arrestor"],
    ["change over switch"],
    ["domestic casserole"],
    ["bain marie"],
    ["wet grinder"],
    ["dough kneader"],
    ["commercial mixer"],
    ["vegetable cutter"],
    ["electric milk boiler"],
    ["mild steel lpg barbecues"],
    ["large compartmental stainless steel tiffin"],
    ["new lpg cooking appliances"],
    ["tandoor"],
    ["air curtain"],
    ["rice boiler"],
    ["chapati warmer"],
    ["roti making machine"],
    ["semi automatic"],
    ["meat cutting machine"],
    ["idli steamer"],
    ["ss", "thermos"],
    ["solar", "street light"],
    ["solar", "pv panel"],
    ["solar ", "plant"],
    ["solar", "battery"],
    ["solar", "inverter"],
    ["solar", "tublar batteries"],
    ["solar ", "street light "],
    ["solar", "water heater"],
    ["solar", "water pump"],
    ["water treament"],
    ["ro plants"],
    ["stp"],
    ["wtp"],
    ["security surveillance"],
    ["cctv"],
    ["body worn camera"],
    ["anti climb fence"],
    ["electric fence"],
    ["chainlink fence"],
    ["picket steel"],
    ["barbed wire"],
    ["punched tape concertina coil ptcc"],
    ["uav"],
    ["nano uav"],
    ["anti drone system"],
    ["high intensity light infrared beam"],
    ["handheld gps"],
    ["convex security mirror"],
    ["hand held thermal imager"],
    ["weapon sites"],
    ["pnvg"],
    ["lorros"],
    ["clip on weapon sites"],
    ["multi function laser aiming system"],
    ["tactical items"],
    ["miltary rain poncho"],
    ["ghilly suit"],
    ["jungle boots"],
    ["rucksack bags"],
    ["3d multi spectral camo vehicle cover"],
    ["shooting range"],
    ["weapon support system"],
    ["long range acoustic hailing device"],
    ["3d multi spectral camo dress"],
    ["bola wrap remote restrain device"],
    ["material"], 
     ["construction equipment"],
    ["jcb bacholoader"],
    ["skid steer loader"],
    ["cranes"],
    ["forklifts"],
    ["swachh bharat items"],
    ["waste management plants"],
    ["waste management"],
    ["road sweeping machines"],
    ["sewer suction machines"],
    ["dustbin"],
    ["frp"],
    ["commercial vaccum cleaner"],
    ["sanitary napkins incinetator machine ", "smoke controlunit"],
    ["sanitary napkin vending machine"],
    ["energy solutions"],
    ["dg sets"],
    ["automobile batteries other batteries"],
    ["fuel cell genrators"],
    ["inverters"],
    
    ["gym"],
    ["outdoor gym"],
    ["computer"], 
    ["peripherals"],
    ["data management solutions"],
    ["access control solutions"],
    ["cyber security solutions"],
    ["video survelliance ", "analytics solutions"]
]


# Start Excel writer
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.index = df.index + 1
    df = df.sort_index()

    # Write full data
    df.to_excel(writer, sheet_name="All Tenders", index=False, startrow=1)

    # Grouped keyword sheets
    for group in keyword_groups:
        filtered_df = df[df["Item Description"].str.lower().apply(
            lambda text: all(word.lower() in text for word in group) if isinstance(text, str) else False
        )]

        if not filtered_df.empty:
            sheet_title = "+ ".join(word.title() for word in group)[:31]
            filtered_df.to_excel(writer, sheet_name=sheet_title, index=False, startrow=1)

# Excel formatting function
def style_worksheet(ws, title):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    centered = Alignment(wrap_text=True, horizontal='center', vertical='center')
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{title} – {current_date}"
    title_cell.font = Font(size=36, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    ws.print_title_rows = '1:2'
    ws.page_setup.fitToWidth = 1

    header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
    header_font = Font(bold=True, size=20)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = centered

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for i, cell in enumerate(row):
            cell.font = Font(size=20, bold=True)
            cell.border = thin_border
            cell.alignment = centered

            col_header = ws.cell(row=2, column=i + 1).value
            if col_header == 'Day Left':
                h_col = 'E'
                i_col = 'F'
                formula = f'=IF((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW() <= 0, "CLOSED", INT((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW()) & "days")'
                cell.value = formula

    for i, cell in enumerate(ws[2], start=1):
        col_letter = get_column_letter(i)
        col_title = cell.value
        if col_title == 'Qty':
            ws.column_dimensions[col_letter].width = 13
        elif col_title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
            ws.column_dimensions[col_letter].width = 18
        elif col_title == 'Item Description':
            ws.column_dimensions[col_letter].width = 35
        elif col_title == 'Address':
            ws.column_dimensions[col_letter].width = 36
        else:
            ws.column_dimensions[col_letter].width = 18

# Apply formatting to all sheets
wb = load_workbook(output_file)
for sheetname in wb.sheetnames:
    sheet_title = sheetname.replace('_', ' ')
    ws = wb[sheetname]
    style_worksheet(ws, sheet_title)

wb.save(output_file)
print(f"✅ Exported and grouped successfully: {output_file}")






