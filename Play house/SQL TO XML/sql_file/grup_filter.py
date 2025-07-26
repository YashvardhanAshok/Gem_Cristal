import os
import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Your keyword list (shortened here, replace with your full list)
keywords_list = ['Solar Battery cells', '3D Multi Spectral Camo Vehicle Cover', '3D Printer', '3d Multi Spectral Camo Dress', 'A.C Static Meter', 'ALL Types of commercial Gym Equipment', 'AMC OF COMMERCIAL KITCHEN EQUIPMENT', 'AMC OF Gym EQUIPMENT', 'Ac static watthour meters-energy meter', 'Access Control Solutions', 'Air Freight Shipping', 'Air curtain', 'All Range Hospital Furniture', 'All Types of Commercial RO PLANTS', 'All Types of Wire and Cables',"Amc", 'Amc Of Ac', 'Amc Of Commercial Kitchen', 'Amc Of Fire Extinguishers', 'Amc Of Generators', 'Amc Of Gym Equipement', 'Amc Of Kitchen Equipement', 'Amc Of Lightning Arrestors', 'Amc Of Ro And IRP', 'Amc Of Solar Power Plant', 'Amc Of Solar Water Heaters', 'Amc Of Transformers', 'Amc of DG Sets and Transformer', 'AntI Drone system', 'Anti climb Fence', 'Automobile Batteries other batteries', 'Bain Marie', 'Bain marie', 'Barbed Wire', 'Battery', 'Body Worn Camera', 'Bola wrap Remote Restrain device', 'Braille Embosser', 'Bricks', 'Bucket Mop Wringer Trolly', 'Butter', 'CCTV', 'CEW',' Conducted Electrical Weapon', 'CGI Sheet', 'Cement','Chainlink Fence', 'Change over Switch', 'Chapati Warmer', 'Clip On Weapon Sites', 'Commercial Mixer', 'Commercial Vaccum Cleaner', 'Computer and peripherals', 'Construction Of Admin Blocks', 'Construction Of Hospital', 'Construction Of Internal Roads', 'Construction Of Klps For Defense', 'Convex Security Mirror', 'Cranes', 'Cyber Forensics Software', 'Cyber Security Solutions', 'DG SETS', 'Data Management solutions', 'Decorative Bollard', 'Decorative Street Light', 'Development Of Infrastructure For Defense', 'Development Of Sewerage Treatement Plant', 'Development Of Water Supply', 'Domestic casserole', 'Dough Kneader', 'Dough kneader 15kg', 'Dry Ration', 'Rice' , 'Pulses' , 'Sugar' , 'Coffee', 'Tea', 'Dustbin', 'Electric Fence', 'Electric Wires','Cable', 'Electric milk boiler', 'FRP', 'FRP Tank', 'Flood Light', 'Flooring', 'Forklifts', 'Fresh Fruits', 'Fresh Vegetable', 'Fuel Cell', 'Fuel cell genrators', 'GPS', 'GPS', 'Global Positioning System', 'Ghillie Suits', 'Ghilly Suit', 'Gi Pipe','Gyser', 'HHTI (Hand Held Thermal Imagers)', 'Hand Held Gas Detector', 'Hand held Thermal Imager', 'Handheld GPS', 'Hardware Item', 'Headphones', 'High Intensity Light Infrared beam', 'Honey Sucker ',' Sewer Cum Jetting Machine', 'Hybrid UPS', 'Idli Steamer', 'Incinerators', 'Inflatable Shelters', 'Inverters', 'JCB Bacholoader', 'Jet Spray', 'Jungle Boots', 'Kunda Gadi', 'LGSF Building', 'Large compartmental stainless steel tiffin', 'Led Bulbs', 'Less Lethal Weapons', 'Lighting Arrestor', 'Lightning Arrestor', 'Long Range Acoustic Hailing Device', 'Lorros', 'MCB', 'MCCB', 'Meat Cutting Machine', 'Mild Steel LPG Barbecues', 'Milk', 'Milk Boiler', 'Miltary Rain Poncho', 'Miniature Circuit Breaker Switches', 'Monitor', 'Multi Function Laser Aiming System', 'Nano Uav', 'New lpg cooking appliances', 'Oil', 'Online UPS', 'Outdoor Gym', 'Oven', 'PNVG', 'PPGI Sheets','Patient Bed Fowler', 'Patient Care Mattress', 'Picket Steel', 'Pickup Truck', 'Plotter', 'Plywood', 'Porta Cabin', 'Portable Kitchen', 'Portable houses', 'Poultry Product', 'Chicken', 'Egg' , 'Mutton', 'Ppgi Sheet', 'Prefab shelters with puf panel', 'Printer', 'Projector', 'Puff Cabin', 'Puff Shelter', 'Punched Tape concertina Coil PTCC', 'Reverse Osmosis', 'Remote Restraint Device', 'Rice Boiler', 'Rice boiler', 'Road Sweeping Machines', 'Robotics', 'Room Heater', 'Roti Making Machine', 'Roti Making Machine Auto matic', 'Rucksack Bags', 'SANITARY NAPKIN VENDING MACHINE', 'SS', 'SS Thermos', 'STP', 'Sewage Treatment Plants', 'Sand', 'Sanitary Items', 'Sanitary Napkins Incinetator Machine with Smoke ControlUnit', 'Satellite Tracker', 'Sea Food', 'Search Light', 'Sedan ',' SUVS', 'Semi Automatic', 'Sewer Suction Machines', 'Shooting Range', 'Skid steer Loader', 'Software','Software Defined Radio', 'Solar Battery', 'Solar Lantern', 'Solar PV Panel','Solar Panel', 'Solar PV Plant', 'Solar Power Plant', 'Solar Street Light', 'Solar Street Light all Type', 'Solar Tublar Batteries', 'Solar Water Heater', 'Solar inverter', 'Solar water Heater', 'Solar water pump', 'Speakers', 'Street Light', 'Switch fuse unit', 'Tablet', 'Tandoor', 'Tandoor',' Millimeter', 'Tubes', 'UAV', 'Under Water Torch', 'Unmanned Aerial Vehicle', 'Vaccum Cleaner', 'Vegetable Cutter', 'Video Survelliance ',' Analytics Solutions', 'WTP', 'Walkie Talkie', 'Waste Management', 'Waste Management Plants', 'Water Bowser', 'Water Cooling', 'Water Dispenser', 'Water Tanker', 'Weapon Sight', 'Weapon Sites', 'Weapon Support system', 'Wet Grinder', 'Wheel Barrow', 'X-ray Machine', 'XLPE Cables', 'water cooler']

keywords_list = ["Fish"]
# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

query = "SELECT * FROM tender_data WHERE Live = 'Yes' AND Cancel != 'Cancel'"
df = pd.read_sql(query, conn)

# Drop unwanted columns
drop_cols = [
    'id', "Cancel", "Department", 'matches', 'matched_products', "element_put",
    "consignee_reporting", "date_of_search", "updated_at",
    'file_path','branch','Branch', 'link_href', 'Live', "extended", "L1_update", 'status', 'L_Placeholder'
]
df.drop(columns=[col for col in drop_cols if col in df.columns], inplace=True)
df.replace(0, '', inplace=True)

# Format tender value column to readable words
def convert_to_words(val):
    try:
        val = float(val)
        if val >= 1_00_00_000:
            return f"{val / 1_00_00_000:.1f} Cr"
        elif val >= 1_00_000:
            return f"{val / 1_00_000:.1f} LPA"
        elif val > 0:
            return f"{val:.0f}"
        else:
            return ""
    except:
        return ""

tender_value_col = next((col for col in df.columns if col.lower() == "tender_value"), None)
if tender_value_col:
    df['Ten-Val Word'] = df[tender_value_col].apply(convert_to_words)

df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

if 'Tender Valu' in df.columns and 'Ten-Val Word' in df.columns:
    cols = df.columns.tolist()
    cols.remove('Ten-Val Word')
    tender_index = cols.index('Tender Valu')
    cols.insert(tender_index + 1, 'Ten-Val Word')
    df = df[cols]

save_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "xl files"))
os.makedirs(save_dir, exist_ok=True)
output_file = os.path.join(save_dir, "Main_By_Keyword.xlsx")

if os.path.exists(output_file):
    os.remove(output_file)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
centered = Alignment(wrap_text=True, horizontal='center', vertical='center')
header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
bold_font_header = Font(bold=True, size=20)
bold_font_title = Font(size=25, bold=True)
current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

wb = Workbook()
wb.remove(wb.active)
import openpyxl
for keyword in keywords_list:
    keyword_lower = keyword.lower()
    filtered = df[df['Item Description'].str.lower().str.contains(keyword_lower, na=False)]

    if filtered.empty:
        print(f"❌ No match found for '{keyword}'")
        continue
    
    sheet_name = keyword[:31].replace('/', '_').replace('\\', '_')
    ws = wb.create_sheet(title=sheet_name)
    ws.print_title_rows = '1:2'
    max_col = ws.max_column

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.auto_filter.ref = f"G2:{get_column_letter(max_col)}2"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or openpyxl.worksheet.properties.PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.horizontalCentered = False

    ws.page_setup.duplex = 2  # 1 = short-edge, 2 = long-edge

    max_col = filtered.shape[1]

    # Add title merged across the columns (row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{keyword} - {current_date}"
    title_cell.font = bold_font_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Write dataframe headers and data starting at row 2
    rows = dataframe_to_rows(filtered, index=False, header=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Style header row (2nd row)
            if r_idx == 2:
                cell.fill = header_fill
                cell.font = bold_font_header
                cell.border = thin_border
                cell.alignment = centered
            else:
                # Style data rows
                cell.font = Font(size=24, bold=True)
                cell.border = thin_border
                cell.alignment = centered

                # Add formula for 'Day Left' column
                col_name = ws.cell(row=2, column=c_idx).value
                if col_name == 'Day Left':
                    # Adjust these column letters as per your sheet layout (example: E and F)
                    h_col = 'E'
                    i_col = 'F'
                    formula = (
                        f'=IF((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW() <= 0, '
                        f'"CLOSED", INT((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW()) & " days")'
                    )
                    cell.value = formula

    # Set column widths
    for col_idx, cell in enumerate(ws[2], start=1):
        col_letter = get_column_letter(col_idx)
        title = cell.value
        if title == 'Qty':
            ws.column_dimensions[col_letter].width = 13
        elif title in [ 'Day Left']:
            ws.column_dimensions[col_letter].width = 14
        elif title in ['End Time']:
            ws.column_dimensions[col_letter].width = 18
        elif title in ['Start Date', 'End Date' ]:
            ws.column_dimensions[col_letter].width = 22
        elif title == 'Item Description':
            ws.column_dimensions[col_letter].width = 30
        elif title == 'Address':
            ws.column_dimensions[col_letter].width = 36
        elif title == "Item Category":
            ws.column_dimensions[col_letter].width = 0
        else:
            ws.column_dimensions[col_letter].width = 18

# Save workbook
wb.save(output_file)
print(f"✅ Exported data grouped by keywords to {output_file}")
