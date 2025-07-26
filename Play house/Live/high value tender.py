import PyPDF2
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import re
import uuid
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import traceback
import pandas as pd
import json
from datetime import date
import pyodbc
from datetime import datetime
today = date.today()
# import win32com.client as win32
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2._page import PageObject
import configparser
config = configparser.ConfigParser()
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from time import sleep, time 
import tkinter as tk
from tkinter import ttk


def wait_for_user():
    def on_click():
        root.destroy()  

    root = tk.Tk()
    root.title("Resume Script")
    root.geometry("200x80")

    style = ttk.Style()
    style.configure("TButton", font=("Segoe UI", 10))

    ttk.Label(root, text="Click to continue").pack(pady=5)
    ttk.Button(root, text="Proceed", command=on_click).pack()
    root.mainloop()  






def get_driver():
    options = Options()
    profile_path = os.path.join(os.getcwd(), "chrome_profile_palladium")
    options.add_argument(f"user-data-dir={profile_path}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    prefs = {
        "download.default_directory": r"C:\vs_code\TenderHunter2.1.3\final\curent and result\paladiam _d file",
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })

    return driver

def login_if_needed(driver):
    try:
        email_input = driver.find_element(By.ID, "normal_login_email")
        password_input = driver.find_element(By.ID, "normal_login_password")
        email_input.clear()
        email_input.send_keys("nss@crystalworks.in")
        password_input.clear()
        password_input.send_keys("macintosh7436", Keys.ENTER)
        sleep(5)
        print("Logged in.")
    except NoSuchElementException:
        print("Already logged in.")

def clean_string(text):
    text = text.replace('\n', ' ')
    text = re.sub(r'\([^)]*\)', '', text)
    text = re.sub(r'\d+', '', text)
    text = re.sub(r'[^a-zA-Z\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def sql_to_json_exclude_columns():
    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=localhost\\SQLEXPRESS;"
        "DATABASE=gem_tenders;"
        "Trusted_Connection=yes;"
    )

    global gem_id_find
    query = f"SELECT * FROM tender_data WHERE tender_id in ({gem_id_find})"

    df = pd.read_sql(query, conn)
    conn.close()

    cols_to_drop = [
        "link_href","end_time","consignee_reporting", "ministry", "department", "id", "date_of_search",
        "element_put", "item_category", "MSE", "branch", "matches", "matched_products",
        "status", "L_Placeholder", "extended", "Cancel", "L1_update", "updated_at", "day_left_formula"
    ]
    df = df.drop(columns=[col for col in cols_to_drop if col in df.columns], errors='ignore')

    # Convert datetime columns to string (format: YYYY-MM-DD)
    date_cols = ['start_date', 'end_date']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')

    json_data = df.to_json(orient='records', indent=4)
    return json_data

def click_download_and_get_file(driver, download_dir, timeout=120):
    try:
        # Wait for and click the Download button
        download_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Download']]"))
        )
        download_button.click()
    except Exception as e:
        return None

    print("⏳ Waiting for download to complete...")
    start_time = time()
    while time() - start_time < timeout:
        files = os.listdir(download_dir)
        # Look for fully downloaded files (not .crdownload)
        completed_files = [f for f in files if not f.endswith(".crdownload")]
        if completed_files:
            # Return the most recent file
            latest_file = max([os.path.join(download_dir, f) for f in completed_files], key=os.path.getctime)
            print(f"✅ Download completed: {latest_file}")
            return latest_file
        sleep(1)

    return None

import os
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2._page import PageObject
from win32com.client import Dispatch
from time import sleep
from PyPDF2 import PdfReader, PdfWriter, PageObject
from win32com.client import Dispatch
import pythoncom

def pdf_maker(excel_file, your_pdf, merged_pdf):
    try:
        pythoncom.CoInitialize()
        print("Starting Excel application...")
        excel = Dispatch('Excel.Application')

        print(f"Opening Excel file: {excel_file}")
        wb = excel.Workbooks.Open(excel_file)

        # Setup page layout for each sheet
        for sheet in wb.Sheets:
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False

        # Export Excel as PDF
        excel_file_pdf_file = os.path.splitext(excel_file)[0] + ".pdf"
        print(f"Exporting Excel to PDF: {excel_file_pdf_file}")
        wb.ExportAsFixedFormat(0, excel_file_pdf_file)
        wb.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()

        sleep(2)

        # Open the input PDF
        with open(excel_file_pdf_file, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            writer = PyPDF2.PdfWriter()

            num_pages = len(reader.pages)

            for i in range(num_pages):
                page = reader.pages[i]

                # Flip every 2nd page (i.e., page index 1, 3, 5,...)
                if (i + 1) % 2 == 0:
                    page.rotate(180)

                writer.add_page(page)

            # Save the output
            with open(excel_file_pdf_file, "wb") as out_file:
                writer.write(out_file)

        # Merge PDFs
        print(f"Merging PDFs: {your_pdf} + {excel_file_pdf_file}")
        your_reader = PdfReader(your_pdf)
        excel_reader = PdfReader(excel_file_pdf_file)
        writer = PdfWriter()

        # Add pages from your PDF
        for page in your_reader.pages:
            writer.add_page(page)

        # Add blank page if your PDF has odd number of pages
        if len(your_reader.pages) % 2 != 0:
            width = your_reader.pages[0].mediabox.width
            height = your_reader.pages[0].mediabox.height
            blank_page = PageObject.create_blank_page(width=width, height=height)
            writer.add_page(blank_page)

        # Add pages from Excel PDF
        for page in excel_reader.pages:
            writer.add_page(page)

        # Write merged PDF
        with open(merged_pdf, "wb") as f:
            writer.write(f)
        print(f"✅ Merged PDF created at: {merged_pdf}")

        # Truncate merged PDF to first 52 pages
        merged_reader = PdfReader(merged_pdf)
        final_writer = PdfWriter()

        for i, page in enumerate(merged_reader.pages):
            if i < 52:
                final_writer.add_page(page)
            else:
                break

        # Save truncated PDF (overwrite merged_pdf or use another file)
        with open(merged_pdf, "wb") as f:
            final_writer.write(f)
        print(f"✅ Final PDF (1–52 pages only) saved at: {merged_pdf}")

    except Exception as e:
        # print("❌ Error in pdf_maker:", e)
        traceback.print_exc() 
        
    finally:
        try:
            excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

from urllib.parse import urlparse
def exl(existing_data,csv_path):
    tender_file = existing_data[0]
    your_pdf = tender_file["file_path"]

    def extract_bid_number(url):
        parsed = urlparse(url)
        if parsed.netloc == "bidplus.gem.gov.in":
            match = re.search(r'(\d+)(?!.*\d)', parsed.path)
            if match:
                return match.group(1)
        return None

    bid_number = extract_bid_number(your_pdf)
    if bid_number:
        your_pdf = f'C:\\vs_code\\TenderHunter2.1.3\\download_pdf\\GeM-Bidding-{bid_number}.pdf'
    
    tender_id = tender_file["tender_id"].replace("/", "_")
    df_json = pd.DataFrame(existing_data)

    df_csv = pd.read_csv(csv_path)
    df_csv = df_csv.drop(columns=['region', 'tender_id', 'url'], errors='ignore')
    df_csv = df_csv[df_csv['bid_rank'].astype(str).str.strip().isin(["L1"])]
    df_csv = df_csv[df_csv['result_bid_value'].astype(str).str.strip() != ""]
     
    df_csv = df_csv[~df_csv['organisation'].astype(str).str.lower().str.contains("ministry")]
    df_csv = df_csv[~df_csv['organisation'].astype(str).str.lower().str.contains("departmen")]
    df_csv.loc[
        df_csv['bid_amount'].astype(str).str.strip() == df_csv['result_bid_value'].astype(str).str.strip(),
            'bid_amount'
    ] = ""
    
    for col in df_csv.select_dtypes(include=['object']).columns:
        df_csv[col] = df_csv[col].str.replace('[', '', regex=False).str.replace(']', '', regex=False)

    df_csv = df_csv.rename(columns={
        'title': 'item_description',
        'ref_no': 'tender_id',
        'status_update_date': 'start_date',
        'state': 'address',
        'bid_amount': 'tender_value'
    })

    df = pd.concat([df_json, df_csv], ignore_index=True)
    df = df.replace(0, '')  # Replace zero values with empty string

    def convert_to_words(val):
        try:
            val = float(val)
            if val >= 1_00_00_000:
                return f"{val / 1_00_00_000:.1f} Cr"
            elif val >= 1_00_000:
                return f"{val / 1_00_000:.1f} L"
            elif val > 0:
                return f"{val:.0f}"
            else:
                return ""
        except:
            return ""

    if "result_bid_value" in df.columns:
        df['RE-Val Word'] = df['result_bid_value'].apply(convert_to_words)

    df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]
    cols = df.columns.tolist()
    if 'Result Bid Value' in cols and 'RE-Val Word' in cols:
        idx = cols.index('Result Bid Value')
        cols.remove('RE-Val Word')
        cols.insert(idx + 1, 'RE-Val Word')
        df = df[cols]

    output_file = r"C:\vs_code\TenderHunter2.1.3\final\Exported_Tender_Data.xlsx"
    sheet_name = ognisation
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.index = df.index + 1
        df = df.sort_index()
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    cols_to_delete = []
    for col_idx, col_cell in enumerate(ws[2], 1):  # header row 2
        if col_cell.value in ['Qty', 'File Path']:
            cols_to_delete.append(col_idx)

    for col_idx in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col_idx)

    # --- Page setup and print titles ---
    ws.print_title_rows = '1:2'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{sheet_name} – Exported on {current_date}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Header row styling ---
    header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(size=20, bold=True)
        cell.border = thin_border
        cell.alignment = alignment

    # --- Body rows styling ---
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for idx, cell in enumerate(row):
            cell.font = Font(size=20, bold=True)
            cell.border = thin_border
            cell.alignment = alignment

    # --- Set column widths ---
    for col_idx, col_cell in enumerate(ws[2], 1):
        col_letter = get_column_letter(col_idx)
        title = col_cell.value
        if title == 'Start Date' or title == 'End Date' or title == 'End Time' or title == 'Day Left':
            ws.column_dimensions[col_letter].width = 18
        elif title == 'Item Description':
            ws.column_dimensions[col_letter].width = 35
        elif title == 'Address':
            ws.column_dimensions[col_letter].width = 20
        elif title == 'Organisation':
            ws.column_dimensions[col_letter].width = 36
        elif title == 'Company Name':
            ws.column_dimensions[col_letter].width = 36
        else:
            ws.column_dimensions[col_letter].width = 18

    wb.save(output_file)
    os.remove(csv_path)

    unique_id = uuid.uuid4()
    print(unique_id)
    merged_pdf = f"C:\\vs_code\\TenderHunter2.1.3\\final\\curent and result\\output_pdf\\{tender_id}_{ognisation}.pdf" 
    print("\n","\n","\n","\n","\n","\n","\n","\n","\n",output_file,"\n",your_pdf,"\n",merged_pdf)
    pdf_maker(output_file,your_pdf,merged_pdf)
    


def main():
    download_dir =  r"C:\vs_code\TenderHunter2.1.3\final\curent and result\paladiam _d file"
    os.makedirs(download_dir, exist_ok=True)

    driver = get_driver()
    driver.get("https://app.palladium.primenumbers.in/results")
    
    login_if_needed(driver)
    sleep(5)

    # sleep(1000)

    tenders_json = sql_to_json_exclude_columns()
    tenders = json.loads(tenders_json)
    print(tenders)
    for tender in tenders: 
        try:
            search_input = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "rc_select_3"))
            )

            search_input.click()
            search_input.send_keys(Keys.CONTROL, 'a')
            search_input.send_keys(Keys.BACKSPACE)
            sleep(0.2)

            input_str = clean_string(str(tender.get("item_description", "")))
            print("\n", input_str)
            search_input.send_keys(input_str)
            sleep(1)

            search_btn = driver.find_element(By.XPATH, "//button[contains(@class, 'ant-input-search-button')]")
            search_btn.click()
            sleep(3)
            wait_for_user()

            # Wait for the checkbox to be present and clickable
            try:
                checkbox_label = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//label[contains(@class, "custom-facet-checkbox")][contains(.,"GEM")]'))
                )
                checkbox_label.click()
            except:
                wait_for_user()
            

            no_result_elements = driver.find_elements(By.XPATH, "//h5[contains(@class, 'fw-bold') and contains(text(), 'No Results Found')]")
            if no_result_elements:
                wait_for_user()

            downloaded_file = click_download_and_get_file(driver, download_dir)
            if downloaded_file:
                tender_arra=[]
                tender["stage"]= "LIVE"
                
                tender_arra.append(tender)
                exl(tender_arra,str(downloaded_file))

        except Exception as e:
            traceback.print_exc()



ognisation = "NDRF_16_7_2025"

# gem_id_find = ["GEM/2025/B/6423449","GEM/2025/B/6434712","GEM/2025/B/6267035","GEM/2025/B/6301369","GEM/2025/B/6259584",]
# gem_id_find = ["GEM/2025/B/6428166","GEM/2025/B/6423682","GEM/2025/B/6249052","GEM/2025/B/6262367","GEM/2025/B/6185838","GEM/2025/B/6241528","GEM/2025/B/6291845","GEM/2025/B/6335486",]
# gem_id_find = ["GEM/2025/B/6392956","GEM/2025/B/6372842","GEM/2025/B/6318209","GEM/2025/B/6305040","GEM/2025/B/6305481"]
gem_id_find = ['GEM/2025/B/6472026']




gem_id_find = ",".join(f"'{tid}'" for tid in gem_id_find)
main()









        





































# Step 1: Load existing JSON-like data
