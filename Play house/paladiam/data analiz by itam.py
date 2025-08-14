import PyPDF2
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import re
import uuid
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import traceback
import pandas as pd
import json
from datetime import date
import pyodbc
from datetime import datetime
today = date.today()
import PyPDF2

from PyPDF2 import PdfReader, PdfWriter
from PyPDF2._page import PageObject
import configparser
config = configparser.ConfigParser()
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from time import sleep, time 
import tkinter as tk
from tkinter import ttk
import os
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2._page import PageObject
from win32com.client import Dispatch
from time import sleep
from PyPDF2 import PdfReader, PdfWriter, PageObject
from win32com.client import Dispatch
import pythoncom

def wait_for_user():
    def on_click():
        root.destroy()  

    root = tk.Tk()
    root.title("Resume Script")
    root.geometry("200x80")

    style = ttk.Style()
    style.configure("TButton", font=("Segoe UI", 10))

    ttk.Label(root, text="Click to continue").pack(pady=5)
    ttk.Button(root, text="Proceed", command=on_click).pack()
    root.mainloop()  

def get_driver():
    options = Options()
    profile_path = os.path.join(os.getcwd(), "chrome_profile_palladium")
    options.add_argument(f"user-data-dir={profile_path}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    prefs = {
        "download.default_directory": r"C:\vs_code\TenderHunter2.1.3\Play house\paladiam\paladiam _d file",
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })

    return driver

def login_if_needed(driver):
    try:
        email_input = driver.find_element(By.ID, "normal_login_email")
        password_input = driver.find_element(By.ID, "normal_login_password")
        email_input.clear()
        email_input.send_keys("nss@crystalworks.in")
        password_input.clear()
        password_input.send_keys("macintosh7436", Keys.ENTER)
        sleep(5)
        print("Logged in.")
    except NoSuchElementException:
        print("Already logged in.")

def clean_string(text):
    text = text.replace('\n', ' ')
    text = re.sub(r'\([^)]*\)', '', text)
    text = re.sub(r'\d+', '', text)
    text = re.sub(r'[^a-zA-Z\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def sql_to_json_exclude_columns():
    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=localhost\\SQLEXPRESS;"
        "DATABASE=gem_tenders;"
        "Trusted_Connection=yes;"
    )

    global gem_id_find
    query = f"SELECT * FROM tender_data WHERE tender_id in ({gem_id_find})"

    df = pd.read_sql(query, conn)
    conn.close()

    cols_to_drop = [
        "epbg_percentage", "link_href","end_time","consignee_reporting", "ministry", "department", "id", "date_of_search",
        "element_put", "item_category", "MSE", "branch", "matches", "matched_products",
        "status", "L_Placeholder", "extended", "Cancel", "L1_update", "updated_at", "day_left_formula","state" 
    ]
    df = df.drop(columns=[col for col in cols_to_drop if col in df.columns], errors='ignore')

    # Convert datetime columns to string (format: YYYY-MM-DD)
    date_cols = ['start_date', 'end_date']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')

    json_data = df.to_json(orient='records', indent=4)
    return json_data

def click_download_and_get_file(driver, download_dir, timeout=120):
    try:
        # Wait for and click the Download button
        download_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Download']]"))
        )
        download_button.click()
    except Exception as e:
        return None

    print("⏳ Waiting for download to complete...")
    start_time = time()
    while time() - start_time < timeout:
        files = os.listdir(download_dir)
        completed_files = [f for f in files if not f.endswith(".crdownload")]
        if completed_files:
            latest_file = max([os.path.join(download_dir, f) for f in completed_files], key=os.path.getctime)
            print(f"✅ Download completed: {latest_file}")
            return latest_file
        sleep(1)

    return None

from urllib.parse import urlparse
def exl(existing_data, csv_path):
    tender_file = existing_data[0]
    tender_id = tender_file["tender_id"].replace("/", "_")

    # Load existing + CSV
    df_json = pd.DataFrame(existing_data)
    df_csv = pd.read_csv(csv_path)
    df_csv = df_csv.drop(columns=['region', 'tender_id', 'url'], errors='ignore')
    df_csv = df_csv[df_csv['bid_rank'].astype(str).str.strip().isin(["L1"])]
    df_csv = df_csv[df_csv['result_bid_value'].astype(str).str.strip() != ""]

    df_csv.loc[
        df_csv['bid_amount'].astype(str).str.strip() ==
        df_csv['result_bid_value'].astype(str).str.strip(),
        'bid_amount'
    ] = ""

    # Safely clean brackets for object columns only
    for col in df_csv.select_dtypes(include=['object']).columns:
        df_csv[col] = (
            df_csv[col]
            .fillna('')   # avoid NaN issues
            .astype(str)
            .str.replace('[', '', regex=False)
            .str.replace(']', '', regex=False)
        )

    # Rename columns
    df_csv = df_csv.rename(columns={
        'title': 'item_description',
        'ref_no': 'tender_id',
        'status_update_date': 'start_date',
        'state': 'address',
        'bid_amount': 'tender_value'
    })

    # Convert to date
    df_csv['start_date'] = pd.to_datetime(df_csv['start_date'], errors='coerce').dt.date
    df_csv = df_csv.dropna(subset=['start_date'])

    # Custom financial year: 1 April - 31 March
    def get_financial_year(date):
        if date.month >= 4:
            return f"FY {date.year}-{str(date.year + 1)[-2:]}"
        else:
            return f"FY {date.year - 1}-{str(date.year)[-2:]}"

    df_csv['financial_year'] = df_csv['start_date'].apply(get_financial_year)

    # Sort by FY and date
    df_csv = df_csv.sort_values(by=['financial_year', 'start_date'], ascending=[False, False]).reset_index(drop=True)

    # Insert year headers
    fy_rows = []
    for fy in sorted(df_csv['financial_year'].unique(), reverse=True):
        fy_rows.append({'item_description': fy, 'financial_year': fy})
        fy_rows.extend(df_csv[df_csv['financial_year'] == fy].to_dict(orient='records'))

    df = pd.DataFrame(fy_rows).reset_index(drop=True)

    # Append to existing JSON data
    df = pd.concat([df_json, df], ignore_index=True)
    df = df.replace(0, '')

    # Convert numbers to human-readable words
    def convert_to_words(val):
        try:
            val = float(val)
            if val >= 1_00_00_000:
                return f"{val / 1_00_00_000:.1f} Cr"
            elif val >= 1_00_000:
                return f"{val / 1_00_000:.1f} L"
            elif val > 0:
                return f"{val:.0f}"
            else:
                return ""
        except:
            return ""

    if "result_bid_value" in df.columns:
        df['RE-Val Word'] = df['result_bid_value'].apply(convert_to_words)

    # Format column headers
    df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

    # Move RE-Val Word next to Result Bid Value
    cols = df.columns.tolist()
    if 'Result Bid Value' in cols and 'RE-Val Word' in cols:
        idx = cols.index('Result Bid Value')
        cols.remove('RE-Val Word')
        cols.insert(idx + 1, 'RE-Val Word')
        df = df[cols]

    # Export to Excel
    output_file = fr"C:\vs_code\TenderHunter2.1.3\{item}_Tender_Data.xlsx"
    sheet_name = item
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.index = df.index + 1
        df = df.sort_index()
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    # Load workbook for styling
    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Delete unwanted columns
    cols_to_delete = []
    for col_idx, col_cell in enumerate(ws[2], 1):
        if col_cell.value in ['Qty', 'File Path']:
            cols_to_delete.append(col_idx)
    for col_idx in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col_idx)

    # Page setup
    ws.print_title_rows = '1:2'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    # Title row
    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{sheet_name} – Exported on {current_date}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Header styling
    header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(size=20, bold=True)
        cell.border = thin_border
        cell.alignment = alignment

    # Body styling
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(size=20, bold=True)
            cell.border = thin_border
            cell.alignment = alignment

    # Column widths
    for col_idx, col_cell in enumerate(ws[2], 1):
        col_letter = get_column_letter(col_idx)
        title = col_cell.value
        if title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
            ws.column_dimensions[col_letter].width = 18
        elif title == 'Item Description':
            ws.column_dimensions[col_letter].width = 35
        elif title == 'Address':
            ws.column_dimensions[col_letter].width = 20
        elif title == 'Organisation':
            ws.column_dimensions[col_letter].width = 36
        elif title == 'Company Name':
            ws.column_dimensions[col_letter].width = 36
        else:
            ws.column_dimensions[col_letter].width = 18

    wb.save(output_file)
    os.remove(csv_path)  
    print(uuid.uuid4())
    print("\n" * 5, output_file )

def main():
    download_dir =  r"C:\vs_code\TenderHunter2.1.3\Play house\paladiam\paladiam _d file"
    os.makedirs(download_dir, exist_ok=True)

    driver = get_driver()
    driver.get("https://app.palladium.primenumbers.in/results")
    
    login_if_needed(driver)
    sleep(5)

    # sleep(1000)

    tenders_json = sql_to_json_exclude_columns()
    tenders = json.loads(tenders_json)
    print(tenders)
    for tender in tenders: 
        try:
            search_input = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "rc_select_3"))
            )

            search_input.click()
            search_input.send_keys(Keys.CONTROL, 'a')
            search_input.send_keys(Keys.BACKSPACE)
            sleep(0.2)

            input_str = clean_string(item)
            print("\n", input_str)
            search_input.send_keys(input_str)
            sleep(1)

            search_btn = driver.find_element(By.XPATH, "//button[contains(@class, 'ant-input-search-button')]")
            search_btn.click()
            sleep(3)
            wait_for_user()

            try:
                checkbox_label = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//label[contains(@class, "custom-facet-checkbox")][contains(.,"GEM")]'))
                )
                checkbox_label.click()
            except:
                wait_for_user()

            no_result_elements = driver.find_elements(By.XPATH, "//h5[contains(@class, 'fw-bold') and contains(text(), 'No Results Found')]")
            if no_result_elements:
                wait_for_user()

            downloaded_file = click_download_and_get_file(driver, download_dir)
            if downloaded_file:
                tender_arra=[]
                tender["stage"]= "LIVE"
                
                tender_arra.append(tender)
                exl(tender_arra,str(downloaded_file))

        except Exception as e:
            traceback.print_exc()

item = "high pressure portable pump for large fire fighting as per 12717"
gem_id_find = ['GEM/2025/B/6530436']

gem_id_find = ",".join(f"'{tid}'" for tid in gem_id_find)
main()
