import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

# Fetch data
query = "SELECT * FROM tender_data"
df = pd.read_sql(query, conn)

# Columns to remove
columns_to_drop = ['id', 'matches', 'matched_products', "element_put", "consignee_reporting", "item_category"]
df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True)

# Replace 0s with empty strings
df = df.replace(0, '')

# Identify the department column
department_col = next((col for col in df.columns if col.strip().lower() == "department"), None)

# Rename columns
df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

# Output file
output_file = "styled_tender_export.xlsx"

if department_col:
    department_col = department_col.replace('_', ' ').title()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sheet_names = []

        for department, group_df in df.groupby(department_col):
            dept_name = str(department).strip() if pd.notna(department) and str(department).strip() else "MINISTRY OF COMMUNICATIONS"
            safe_sheet_name = dept_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_') or "MINISTRY OF COMMUNICATIONS"
            sheet_names.append(safe_sheet_name)

            # Sort by Start Date (latest first)
            if 'Start Date' in group_df.columns:
                group_df['Start Date'] = pd.to_datetime(group_df['Start Date'], errors='coerce')
                group_df = group_df.sort_values('Start Date', ascending=False)

            # Filter out rows that are marked as 'CLOSED' (temporarily if already computed)
            if 'Day Left' in group_df.columns:
                group_df = group_df[~group_df['Day Left'].astype(str).str.upper().str.contains('CLOSED')]

            # Reset index after filtering
            group_df.index = range(len(group_df))

            # Write data to Excel starting from row 2
            group_df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=1)

    # Open the workbook
    wb = load_workbook(output_file)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    centered_wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # Add title to first row
        max_col = ws.max_column
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=1, column=1).value = f"{sheet_name} – Exported on {current_date}"
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        # Style header row (row 2)
        header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
        bold_font = Font(bold=True, size=15)
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = centered_wrap_alignment

        # Apply auto-filter from header (row 2)
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

        # Style data rows (starting from row 3)
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
            ws.row_dimensions[row_idx].height = 120
            for idx, cell in enumerate(row):
                cell.font = Font(size=15)
                cell.border = thin_border
                cell.alignment = centered_wrap_alignment

                col_name = ws.cell(row=2, column=idx + 1).value
                if col_name == 'Day Left':
                    start_date_col = get_column_letter(idx - 2)
                    end_date_col = get_column_letter(idx - 1)
                    formula = f'=IF(({start_date_col}{row_idx}+{end_date_col}{row_idx})-NOW() <= 0, "CLOSED", INT(({start_date_col}{row_idx}+{end_date_col}{row_idx})-NOW()) & " days")'
                    cell.value = formula
                    cell.font = Font(size=18, color="FF0000")

        # Set column widths
        for col_idx, col_cell in enumerate(ws[2], start=1):
            col_letter = get_column_letter(col_idx)
            col_title = col_cell.value
            if col_title == 'Qty':
                ws.column_dimensions[col_letter].width = 7
            elif col_title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
                ws.column_dimensions[col_letter].width = 13
            elif col_title == 'Item Description':
                ws.column_dimensions[col_letter].width = 35
            elif col_title == 'Address':
                ws.column_dimensions[col_letter].width = 40
            else:
                ws.column_dimensions[col_letter].width = 18

        # Page Setup: Landscape and Narrow margins
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25

    wb.save(output_file)
    print(f"✅ Data exported successfully to {output_file} with header in row 2, filters, and formatting applied.")
