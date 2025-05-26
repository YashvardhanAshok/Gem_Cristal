import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.page import PageMargins

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

query = "SELECT * FROM tender_data"
df = pd.read_sql(query, conn)

columns_to_drop = ['id', 'matches', 'matched_products', "element_put",  "consignee_reporting", "DATE OF SEARCH", "link"]
df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True)

keywords = ['Solar',"AMC"]

def filter_rows(row):
    desc = str(row.get("item_description", "")).lower()
    cat = str(row.get("item_category", "")).lower()
    return (
        all(keyword.lower() in desc for keyword in keywords) or
        all(keyword.lower() in cat for keyword in keywords)
    )


# Apply filter
df = df[df.apply(filter_rows, axis=1)]
df = df.replace(0, '')
df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]
output_file = "styled_tender_filtered_export.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.index = df.index + 1  # to leave a blank row at top
    df = df.sort_index()
    df.to_excel(writer, sheet_name="Filtered Data", index=False, startrow=1)

wb = load_workbook(output_file)
ws = wb["Filtered Data"]
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
centered_wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

# Add header row
sheet_title = f"Filtered Export – {current_date}"
max_col = ws.max_column
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
title_cell = ws.cell(row=1, column=1)
title_cell.value = sheet_title
title_cell.font = Font(size=16, bold=True)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Autofilter
ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"

# Page settings
ws.print_title_rows = '1:2'
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)

# Header row style
header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
bold_font = Font(bold=True, size=15)
for cell in ws[2]:
    cell.fill = header_fill
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = centered_wrap_alignment

# Data rows style
for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    ws.row_dimensions[row_idx].height = 120
    for idx, cell in enumerate(row):
        cell.font = Font(size=15)
        cell.border = thin_border
        cell.alignment = centered_wrap_alignment

        col_name = ws.cell(row=2, column=idx + 1).value
        if col_name == 'Day Left':
            h_col = 'F'  # Start Date
            i_col = 'G'  # End Date
            formula = f'=IF((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW() <= 0, "CLOSED", INT((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW()) & " days")'
            cell.value = formula
            cell.font = Font(size=18, color="FF0000")

# Adjust column widths
for col_idx, col_cell in enumerate(ws[2], start=1):
    col_letter = get_column_letter(col_idx)
    col_title = col_cell.value
    if col_title == 'Qty':
        ws.column_dimensions[col_letter].width = 10
    elif col_title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
        ws.column_dimensions[col_letter].width = 15
    elif col_title == 'Item Description':
        ws.column_dimensions[col_letter].width = 35
    elif col_title == 'Address':
        ws.column_dimensions[col_letter].width = 40
    else:
        ws.column_dimensions[col_letter].width = 18

wb.save(output_file)
print(f"✅ Filtered data exported successfully to {output_file} with all formatting applied.")
