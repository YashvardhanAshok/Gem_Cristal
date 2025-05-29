import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.page import PageMargins

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

# Fetch data
query = "SELECT * FROM tender_data"
df = pd.read_sql(query, conn)

# Columns to remove
columns_to_drop = ['id', 'matches', 'matched_products', "element_put", "consignee_reporting", "item_category", "date_of_search", "link_herf", 'file_path']
df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])

# Replace 0 and False with empty string
df = df.replace({0: '', False: ''})

# Rename columns
df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

# Output file
output_file = "styled_tender_export.xlsx"

# Write to a single sheet
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.index = df.index + 1  # Shift index to leave first row blank
    df = df.sort_index()
    df.to_excel(writer, sheet_name="All Tenders", index=False, startrow=1)

# Style the Excel sheet
wb = load_workbook(output_file)
ws = wb["All Tenders"]
ws.print_title_rows = '1:2'
current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
max_col = ws.max_column

# Merge top row for title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
title_cell = ws.cell(row=1, column=1)
title_cell.value = f"All Tenders – Exported on {current_date}"
title_cell.font = Font(size=16, bold=True)
title_cell.alignment = Alignment(horizontal="left", vertical="center")

# Add autofilter
ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"

# Page setup
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)

# Define styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
bold_font = Font(bold=True, size=15)
cell_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

# Style header
for cell in ws[2]:
    cell.fill = header_fill
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = cell_alignment

# Style rows and insert Day Left formula
for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    ws.row_dimensions[row_idx].height = 120
    for idx, cell in enumerate(row):
        cell.font = Font(size=15)
        cell.border = thin_border
        cell.alignment = cell_alignment

        col_name = ws.cell(row=2, column=idx + 1).value
        if col_name == 'Day Left':
            h_col = 'E'  # Start Date
            i_col = 'F'  # End Date
            formula = f'=IF((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW() <= 0, "CLOSED", INT((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW()) & " days")'
            cell.value = formula
            cell.font = Font(size=18, color="FF0000")

# Set column widths
for col_idx, col_cell in enumerate(ws[2], start=1):
    col_letter = get_column_letter(col_idx)
    col_title = col_cell.value
    if col_title == 'Qty':
        ws.column_dimensions[col_letter].width = 10
    elif col_title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
        ws.column_dimensions[col_letter].width = 15
    elif col_title == 'Item Description':
        ws.column_dimensions[col_letter].width = 35
    elif col_title == 'Address':
        ws.column_dimensions[col_letter].width = 40
    else:
        ws.column_dimensions[col_letter].width = 18

wb.save(output_file)
print(f"✅ Data exported successfully to {output_file} with one sheet and all values included.")
