import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.page import PageMargins
import os

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

# Fetch data
query = "SELECT * FROM tender_data"
df = pd.read_sql(query, conn)

# Columns to remove
columns_to_drop = ['id', 'matches', 'matched_products', "element_put", "consignee_reporting",  "date_of_search", "updated_at", 'file_path', 'link_href', 'Live', "extended", "L1_update", 'status','L_Placeholder']
for col in columns_to_drop:
    if col in df.columns:
        df = df.drop(columns=col)

# Replace 0s with empty strings
df = df.replace(0, '')

# Convert tender_value to readable format
def convert_to_words(val):
    try:
        val = float(val)
        if val >= 1_00_00_000:
            return f"{val / 1_00_00_000:.1f} Cr"
        elif val >= 1_00_000:
            return f"{val / 1_00_000:.1f} LPA"
        elif val > 0:
            return f"{val:.0f}"
        else:
            return ""
    except:
        return ""

# Create new column for formatted tender value
tender_value_col = None
for col in df.columns:
    if col.strip().lower() == "tender_value":
        tender_value_col = col
        break

if tender_value_col:
    df['Ten-Val Word'] = df[tender_value_col].apply(convert_to_words)

# Rename columns for display
df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

# Move 'Ten-Val Word' after 'Tender Valu'
cols = df.columns.tolist()
if 'Tender Valu' in cols and 'Ten-Val Word' in cols:
    tender_index = cols.index('Tender Valu')
    cols.remove('Ten-Val Word')
    cols.insert(tender_index + 1, 'Ten-Val Word')
    df = df[cols]

# Output file path
save_file = os.path.abspath(os.path.join(os.path.dirname(__file__), "xl files")) 
output_file = f"{save_file}/Main.xlsx"

# Write full DataFrame to one sheet
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.index = df.index + 1
    df = df.sort_index()
    df.to_excel(writer, sheet_name="All Tenders", index=False, startrow=1)

# Styling
wb = load_workbook(output_file)
ws = wb["All Tenders"]

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
centered_wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

# Add title row
max_col = ws.max_column
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
title_cell = ws.cell(row=1, column=1)
title_cell.value = f"Tender Export – {current_date}"
title_cell.font = Font(size=16, bold=True)
title_cell.alignment = Alignment(horizontal="left", vertical="center")

# Autofilter
ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"

# Page settings
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_margins = PageMargins(
    left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
)
ws.print_title_rows = '1:2'
ws.page_setup.fitToWidth = 1

# Header style
header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
bold_font = Font(bold=True, size=20)
for cell in ws[2]:
    cell.fill = header_fill
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = centered_wrap_alignment

# Row styling
for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    for idx, cell in enumerate(row):
        cell.font = Font(size=20, bold=True)
        cell.border = thin_border
        cell.alignment = centered_wrap_alignment

        col_name = ws.cell(row=2, column=idx + 1).value
        if col_name == 'Day Left':
            h_col = 'E'
            i_col = 'F'
            formula = f'=IF((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW() <= 0, "CLOSED", INT((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW()) & " days")'
            cell.value = formula

# Adjust column widths
for col_idx, col_cell in enumerate(ws[2], start=1):
    col_letter = get_column_letter(col_idx)
    col_title = col_cell.value
    if col_title == 'Qty':
        ws.column_dimensions[col_letter].width = 13
    elif col_title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
        ws.column_dimensions[col_letter].width = 18
    elif col_title == 'Item Description':
        ws.column_dimensions[col_letter].width = 35
    elif col_title == 'Address':
        ws.column_dimensions[col_letter].width = 36
    else:
        ws.column_dimensions[col_letter].width = 18

wb.save(output_file)
print(f"✅ Exported to single sheet: {output_file}")
