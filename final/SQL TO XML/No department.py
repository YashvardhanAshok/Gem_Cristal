import pyodbc
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.page import PageMargins

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

# Fetch data
query = "SELECT * FROM tender_data"
df = pd.read_sql(query, conn)

# Excel file path
file_path = "tender_data_export.xlsx"

# Save DataFrame to Excel using openpyxl
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Tenders', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Tenders']

    # Apply formatting (optional)
    font_bold = Font(bold=True)
    fill_header = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    # Format header
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = alignment_center
        cell.border = border

    # Autofit columns
    for col_num, col in enumerate(df.columns, 1):
        max_length = max((len(str(value)) for value in df[col]), default=0)
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Optional: Set page margins for printing
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

print(f"Data exported to '{file_path}' successfully.")
