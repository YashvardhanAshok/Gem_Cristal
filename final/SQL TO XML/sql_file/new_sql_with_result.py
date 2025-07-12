import pyodbc
import pandas as pd
import ast
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.page import PageMargins

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

# Fetch data
query = '''
        SELECT * 
        FROM tender_data 
        WHERE date_of_search ='2025-06-30' 
        '''



# query = "SELECT * FROM tender_data where organisation = 'OIL AND NATURAL GAS CORPORATION LIMITED'"
# query = "SELECT * FROM tender_data WHERE department = 'BORDER SECURITY FORCE';"
# query = "SELECT * FROM tender_data WHERE department = 'BORDER SECURITY FORCE';"
# query = "SELECT * FROM tender_data WHERE date_of_search = '2025-05-31';"
df = pd.read_sql(query, conn)

# Columns to remove
columns_to_drop = ['id', 'matches', 'matched_products', "element_put",  "consignee_reporting", "item_category",
                   "date_of_search", "updated_at", 'file_path', 'link_href', 'Live', "extended", "Cancel", "L1_update"]
for col in columns_to_drop:
    if col in df.columns:
        df = df.drop(columns=col)

# Replace 0s with empty strings
df = df.replace(0, '')

# Safely parse and expand L_Placeholder
def expand_l_placeholder(row):
    try:
        values = row.get("L_Placeholder") or row.get("L Placeholder")  # Handle both variations
        if isinstance(values, str):
            values = ast.literal_eval(values)
        if not values:
            return {}

        result = {}
        for i, pair in enumerate(values, start=1):
            if isinstance(pair, list) and len(pair) == 2:
                result[f"L{i}"] = pair[0]
                result[f"L{i} Amount"] = pair[1]
        return result
    except Exception:
        return {}

# Expand the L_Placeholder column
if "L_Placeholder" in df.columns:
    expanded_df = df.apply(expand_l_placeholder, axis=1, result_type='expand')
    df = pd.concat([df.drop(columns=["L_Placeholder"]), expanded_df], axis=1)

# Identify the department column
department_col = None
for col in df.columns:
    if col.strip().lower() == "department":
        department_col = col
        break

# Rename columns
df.columns = [col.replace('_', ' ').title() if col != 'day_left_formula' else 'Day Left' for col in df.columns]

# Output file
import os
save_file = os.path.abspath(os.path.join(os.path.dirname(__file__),"xl files")) 
output_file = f"{save_file}/New main.xlsx"


if department_col:
    department_col = department_col.replace('_', ' ').title()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sheet_names = []

        for department, group_df in df.groupby(department_col):
            dept_name = str(department).strip() if pd.notna(department) and str(department).strip() else "MINISTRY OF COMMUNICATIONS"
            safe_sheet_name = dept_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_')
            if not safe_sheet_name:
                safe_sheet_name = "MINISTRY OF COMMUNICATIONS"
            sheet_names.append(safe_sheet_name)

            # Insert blank row to make space for header
            group_df.index = group_df.index + 1
            group_df = group_df.sort_index()

            group_df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=1)

    # Open the workbook
    wb = load_workbook(output_file)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    centered_wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Skipping: Worksheet '{sheet_name}' does not exist.")
            continue
        ws = wb[sheet_name]


        # Set rows 1 and 2 to repeat on printed pages
        ws.print_title_rows = '1:2'

        # Add sheet title and date at the top row
        sheet_title = f"{sheet_name} – Exported on {current_date}"
        max_col = ws.max_column
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = sheet_title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add autofilter to header row (row 2)
        ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"

        # Set page layout to landscape
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        # Set narrow margins
        ws.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.75,
            bottom=0.75,
            header=0.3,
            footer=0.3
        )

        # Style header row
        header_fill = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")
        bold_font = Font(bold=True, size=15)
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = centered_wrap_alignment

        # Style data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
            ws.row_dimensions[row_idx].height = 120
            for idx, cell in enumerate(row):
                cell.font = Font(size=15)
                cell.border = thin_border
                cell.alignment = centered_wrap_alignment

                col_name = ws.cell(row=2, column=idx + 1).value
                if col_name == 'Day Left':
                    h_col = 'E'  # Start Date
                    i_col = 'F'  # End Date
                    formula = f'=IF((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW() <= 0, "CLOSED", INT((INDIRECT("{h_col}"&ROW())+INDIRECT("{i_col}"&ROW()))-NOW()) & " days")'
                    cell.value = formula
                    cell.font = Font(size=18, color="FF0000")

        # Adjust column widths
        for col_idx, col_cell in enumerate(ws[2], start=1):
            col_letter = get_column_letter(col_idx)
            col_title = col_cell.value
            if col_title == 'Qty':
                ws.column_dimensions[col_letter].width = 10
            elif col_title in ['Start Date', 'End Date', 'End Time', 'Day Left']:
                ws.column_dimensions[col_letter].width = 15
            elif col_title == 'Item Description':
                ws.column_dimensions[col_letter].width = 35
            elif col_title == 'Address':
                ws.column_dimensions[col_letter].width = 40
            else:
                ws.column_dimensions[col_letter].width = 18

    wb.save(output_file)
    print(f"✅ Data exported successfully to {output_file} with header row and all formatting applied.")
else:
    print("❌ Error: 'Department' column not found in the data.")
