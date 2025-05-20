import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Connect to SQL Server
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=gem_tenders;"
    "Trusted_Connection=yes;"
)

query = "SELECT * FROM tender_data"
df = pd.read_sql(query, conn)

id_array = df['tender_id'].tolist()

print("ID Array:", id_array)

total_entries = len(id_array)
print(f"Total entries: {total_entries}")

# 3603